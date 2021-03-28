import os
import openpyxl
import win32com.client as win32
# from openpyxl.styles import Alignment

def index():
    file_list = os.listdir(os.getcwd())
    for name in file_list[:]:
        if not name.endswith(".xls"): file_list.remove(name)
    # print(file_list)
    for name in file_list:
        Xls2Xlsx(os.getcwd() + "/" + name)
        Convert(os.getcwd() + "/" + name + 'x')

def Xls2Xlsx(name):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(name)

    wb.SaveAs(name+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
    wb.Close()                               #FileFormat = 56 is for .xls extension
    excel.Application.Quit()

def Convert(name):
    # 打开工作簿
    wb_raw = openpyxl.load_workbook(name)
    sh_raw = wb_raw['得分明细']
    # 创建新工作簿
    wb = openpyxl.Workbook()
    sh = wb['Sheet']
    offset = 2
    sh.cell(row = 1,column = 1,value = "高三（" + name[name.rfind("班") - 1] + "）班周测")
    sh.merge_cells('A1:N1')
    sh.cell(row = offset,column = 1,value = "学号")
    sh.cell(row = offset,column = 2,value = "姓名")
    sh.cell(row = offset,column = 3,value = "客观分")
    sh.cell(row = offset,column = 4,value = "校次")
    sh.cell(row = offset,column = 5,value = "班次")
    sh.cell(row = offset,column = 6,value = "听力")
    sh.cell(row = offset,column = 7,value = "阅读")
    sh.cell(row = offset,column = 8,value = "七选五")
    sh.cell(row = offset,column = 9,value = "完形")
    sh.cell(row = offset,column = 10,value = "语法")
    sh.cell(row = offset,column = 11,value = "应用文")
    sh.cell(row = offset,column = 12,value = "大作文")
    sh.cell(row = offset,column = 13,value = "总分")
    sh.cell(row = offset,column = 14,value = "排名")
    cnt = offset

    # 按行读取数据，去掉表头和统计
    for stu in list(sh_raw.rows)[2:sh_raw.max_row-4]:
        # 基础信息
        stu_id = int(stu[4].value)
        stu_name = stu[1].value
        stu_rank_school = stu[7].value
        stu_rank_class = stu[8].value
        if stu[9].value != "-" :
            stu_total_score = float(stu[9].value)

            # 分别读取听力、阅读、七选五、完形成绩
            stu_score_listening = 0.0
            for i in range(12,32) :
                stu_score_listening += float(stu[i].value)
            stu_score_reading = 0.0
            for i in range(32,42) :
                stu_score_reading += float(stu[i].value)
            stu_score_choosing5 = 0.0
            for i in range(42,47) :
                stu_score_choosing5 += float(stu[i].value)
            stu_score_closing = 0.0
            for i in range(47,67) :
                stu_score_closing += float(stu[i].value)
        else :
            stu_total_score = "-"
            stu_score_listening = ""
            stu_score_reading = ""
            stu_score_choosing5 = ""
            stu_score_closing = ""

        # print(stu_id,stu_name,stu_total_score,stu_score_listening,stu_score_reading,stu_score_choosing5,stu_score_closing)
        cnt += 1
        sh.cell(row = stu_id + offset,column = 1,value = stu_id)
        sh.cell(row = stu_id + offset,column = 2,value = stu_name)
        sh.cell(row = stu_id + offset,column = 3,value = stu_total_score)
        sh.cell(row = stu_id + offset,column = 4,value = stu_rank_school)
        sh.cell(row = stu_id + offset,column = 5,value = stu_rank_class)
        sh.cell(row = stu_id + offset,column = 6,value = stu_score_listening)
        sh.cell(row = stu_id + offset,column = 7,value = stu_score_reading)
        sh.cell(row = stu_id + offset,column = 8,value = stu_score_choosing5)
        sh.cell(row = stu_id + offset,column = 9,value = stu_score_closing)
        # sh.cell(row = stu_id + offset,column = 10,value = "语法")
        # sh.cell(row = stu_id + offset,column = 11,value = "应用文")
        # sh.cell(row = stu_id + offset,column = 12,value = "大作文")
        sh.cell(row = stu_id + offset,column = 13,value = "=IFERROR(C" + str(stu_id + offset) + "+SUM(J" + str(stu_id + offset) + ":L" + str(stu_id + offset) + "),\"\")")
    
    # 调整格式
    cnt += 1
    for i in range(offset + 1 , cnt):
        sh.cell(row = i,column = 14,value = "=IFERROR(RANK(M" + str(i) + ",$M$" + str(offset + 1) + ":$M$" + str(cnt - 1) + ",0),\"\")")

    # sh['A1:N'+str(cnt)].Alignment = Alignment(horizontal='center', vertical='center')
    sh.cell(row = cnt,column = 1,value = "平均" )
    sh.merge_cells("A" + str(cnt) + ":B" + str(cnt))
    sh.cell(row = cnt,column = 3,value = "=AVERAGE(C" + str(offset + 1) + ":C" + str(cnt - 1) + ")" )
    sh.cell(row = cnt,column = 6,value = "=AVERAGE(F" + str(offset + 1) + ":F" + str(cnt - 1) + ")" )
    sh.cell(row = cnt,column = 7,value = "=AVERAGE(G" + str(offset + 1) + ":G" + str(cnt - 1) + ")" )
    sh.cell(row = cnt,column = 8,value = "=AVERAGE(H" + str(offset + 1) + ":H" + str(cnt - 1) + ")" )
    sh.cell(row = cnt,column = 9,value = "=AVERAGE(I" + str(offset + 1) + ":I" + str(cnt - 1) + ")" )
    sh.cell(row = cnt,column = 10,value = "=AVERAGE(J" + str(offset + 1) + ":J" + str(cnt - 1) + ")" )
    sh.cell(row = cnt,column = 11,value = "=AVERAGE(K" + str(offset + 1) + ":K" + str(cnt - 1) + ")" )
    sh.cell(row = cnt,column = 12,value = "=AVERAGE(L" + str(offset + 1) + ":L" + str(cnt - 1) + ")" )
    sh.cell(row = cnt,column = 13,value = "=AVERAGE(M" + str(offset + 1) + ":M" + str(cnt - 1) + ")" )
    # 保存并关闭工作薄
    wb_raw.close()
    wb.save(os.getcwd() + "/高三（" + name[name.rfind("班") - 1] + "）班周测 00.00.xlsx")
    wb.close()

# def Converts(name):
#     print(name[name.rfind("班") - 1])

if __name__ == '__main__':
    index()
